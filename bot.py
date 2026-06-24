import os
import logging
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    Application, CommandHandler, MessageHandler, CallbackQueryHandler,
    filters, ContextTypes, ConversationHandler
)
from datetime import datetime
import anthropic
from supabase import create_client
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import tempfile

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Environment variables
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_KEY")

# Clients
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
ai_client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)

# Free tier limit
FREE_LIMIT = 20

# Conversation states
(
    WAITING_FOR_RECEIPT,
    CONFIRM_EXPENSE,
    EDIT_MENU,
    EDIT_AMOUNT,
    EDIT_MERCHANT,
    EDIT_WALLET,
    EDIT_DATE,
    MANUAL_AMOUNT,
    MANUAL_MERCHANT,
    MANUAL_WALLET,
    MANUAL_WALLET_DETAIL,
    HISTORY_PICK_DATE,
    EXPORT_PICK_DATE,
) = range(13)

# Main menu keyboard - always visible
MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [[KeyboardButton("➕ Add Expense"), KeyboardButton("📊 My Spending"), KeyboardButton("📤 Export")]],
    resize_keyboard=True,
    is_persistent=True
)

def get_wallet_keyboard(include_back=True):
    buttons = [
        [InlineKeyboardButton("💵 Cash", callback_data="wallet_Cash"),
         InlineKeyboardButton("💚 GCash", callback_data="wallet_GCash")],
        [InlineKeyboardButton("💜 Maya", callback_data="wallet_Maya"),
         InlineKeyboardButton("💳 Credit Card", callback_data="wallet_CreditCard")],
        [InlineKeyboardButton("🏦 Other Banks", callback_data="wallet_OtherBanks")],
    ]
    if include_back:
        buttons.append([InlineKeyboardButton("🔙 Back", callback_data="back_main")])
    return InlineKeyboardMarkup(buttons)

def get_spending_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Today", callback_data="view_today"),
         InlineKeyboardButton("📆 This Week", callback_data="view_week")],
        [InlineKeyboardButton("🗓️ This Month", callback_data="view_month"),
         InlineKeyboardButton("🔍 Pick a Date", callback_data="view_pick")],
        [InlineKeyboardButton("🔙 Back", callback_data="back_main")]
    ])

def get_export_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Today", callback_data="export_today"),
         InlineKeyboardButton("📆 This Week", callback_data="export_week")],
        [InlineKeyboardButton("🗓️ This Month", callback_data="export_month"),
         InlineKeyboardButton("🔍 Pick a Date", callback_data="export_pick")],
        [InlineKeyboardButton("🔙 Back", callback_data="back_main")]
    ])

def get_confirm_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Save", callback_data="confirm_save"),
         InlineKeyboardButton("✏️ Edit", callback_data="confirm_edit"),
         InlineKeyboardButton("❌ Cancel", callback_data="confirm_cancel")]
    ])

def get_edit_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💸 Amount", callback_data="edit_amount"),
         InlineKeyboardButton("🏪 Where", callback_data="edit_merchant")],
        [InlineKeyboardButton("💳 Paid via", callback_data="edit_wallet"),
         InlineKeyboardButton("📅 Date", callback_data="edit_date")],
        [InlineKeyboardButton("🔙 Back", callback_data="back_confirm")]
    ])

# DB helpers
def get_or_create_user(telegram_id, username, first_name):
    result = supabase.table("users").select("*").eq("id", telegram_id).execute()
    if not result.data:
        supabase.table("users").insert({
            "id": telegram_id,
            "username": username,
            "first_name": first_name,
        }).execute()
        return {"id": telegram_id, "is_pro": False, "transaction_count": 0}
    return result.data[0]

def get_user(telegram_id):
    result = supabase.table("users").select("*").eq("id", telegram_id).execute()
    return result.data[0] if result.data else None

def increment_count(telegram_id):
    user = get_user(telegram_id)
    supabase.table("users").update({"transaction_count": user["transaction_count"] + 1}).eq("id", telegram_id).execute()

def check_limit(telegram_id):
    user = get_user(telegram_id)
    if user["is_pro"]:
        return True
    return user["transaction_count"] < FREE_LIMIT

def save_expense(user_id, amount, merchant, wallet, wallet_detail, date, time, entry_type):
    supabase.table("expenses").insert({
        "user_id": user_id,
        "amount": float(amount),
        "merchant": merchant,
        "wallet": wallet,
        "wallet_detail": wallet_detail,
        "date": date,
        "time": time,
        "entry_type": entry_type
    }).execute()
    increment_count(user_id)

def format_wallet_display(wallet, wallet_detail=None):
    icons = {"Cash": "💵", "GCash": "💚", "Maya": "💜", "CreditCard": "💳", "OtherBanks": "🏦"}
    icon = icons.get(wallet, "💳")
    if wallet_detail:
        return f"{icon} {wallet_detail}"
    names = {"Cash": "Cash", "GCash": "GCash", "Maya": "Maya", "CreditCard": "Credit Card", "OtherBanks": "Other Banks"}
    return f"{icon} {names.get(wallet, wallet)}"

def format_expense_confirmation(data):
    wallet_display = format_wallet_display(data.get("wallet"), data.get("wallet_detail"))
    return (
        f"Here's what I found:\n\n"
        f"💸 Amount:   ₱{float(data['amount']):,.2f}\n"
        f"🏪 Where:    {data['merchant']}\n"
        f"💳 Paid via: {wallet_display}\n"
        f"📅 Date:     {data['date']}\n"
        f"🕐 Time:     {data['time']}\n\n"
        f"Is this correct?"
    )

# AI receipt reading
async def read_receipt(image_bytes):
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    response = ai_client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=500,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": "image/jpeg", "data": b64}
                },
                {
                    "type": "text",
                    "text": (
                        "Read this receipt and extract: amount (numbers only), merchant name, "
                        "wallet/payment method (GCash/Maya/Cash/Credit Card/bank name), date (YYYY-MM-DD), time (HH:MM:SS). "
                        "Reply ONLY in this exact format:\n"
                        "AMOUNT: [number]\n"
                        "MERCHANT: [name]\n"
                        "WALLET: [GCash/Maya/Cash/CreditCard/OtherBanks]\n"
                        "WALLET_DETAIL: [specific card or bank name if applicable, else NONE]\n"
                        "DATE: [YYYY-MM-DD or UNKNOWN]\n"
                        "TIME: [HH:MM:SS or UNKNOWN]"
                    )
                }
            ]
        }]
    )
    text = response.content[0].text.strip()
    data = {}
    for line in text.split("\n"):
        if ":" in line:
            key, val = line.split(":", 1)
            data[key.strip()] = val.strip()

    now = datetime.now()
    return {
        "amount": data.get("AMOUNT", "0"),
        "merchant": data.get("MERCHANT", "Unknown"),
        "wallet": data.get("WALLET", "Cash"),
        "wallet_detail": data.get("WALLET_DETAIL") if data.get("WALLET_DETAIL") != "NONE" else None,
        "date": data.get("DATE") if data.get("DATE") != "UNKNOWN" else now.strftime("%Y-%m-%d"),
        "time": data.get("TIME") if data.get("TIME") != "UNKNOWN" else now.strftime("%H:%M:%S"),
    }

# Spending views
def get_today_expenses(user_id):
    today = datetime.now().strftime("%Y-%m-%d")
    result = supabase.table("expenses").select("*").eq("user_id", user_id).eq("date", today).order("time").execute()
    return result.data

def get_week_expenses(user_id):
    from datetime import timedelta
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    result = supabase.table("expenses").select("*").eq("user_id", user_id).gte("date", str(week_start)).lte("date", str(today)).order("date").execute()
    return result.data

def get_month_expenses(user_id):
    today = datetime.now()
    month_start = today.replace(day=1).strftime("%Y-%m-%d")
    result = supabase.table("expenses").select("*").eq("user_id", user_id).gte("date", month_start).lte("date", today.strftime("%Y-%m-%d")).order("date").execute()
    return result.data

def get_date_expenses(user_id, date_str):
    result = supabase.table("expenses").select("*").eq("user_id", user_id).eq("date", date_str).order("time").execute()
    return result.data

def format_expense_list(expenses, title):
    if not expenses:
        return f"{title}\n\nNo expenses recorded. 🎉"

    total = sum(float(e["amount"]) for e in expenses)
    wallet_totals = {}
    lines = [f"{title}\n"]

    for i, e in enumerate(expenses, 1):
        wallet_display = format_wallet_display(e["wallet"], e.get("wallet_detail"))
        time_str = e["time"][:5] if e["time"] else ""
        lines.append(f"{i}. {e['merchant']}  ₱{float(e['amount']):,.2f}  {wallet_display}  {time_str}")
        wkey = e.get("wallet_detail") or e["wallet"]
        wallet_totals[wkey] = wallet_totals.get(wkey, 0) + float(e["amount"])

    lines.append("─────────────────────────")
    lines.append(f"💸 Total:  ₱{total:,.2f}\n")
    for w, amt in wallet_totals.items():
        icon = format_wallet_display(w).split()[0] if len(w) < 10 else "🏦"
        lines.append(f"{icon} {w}:  ₱{amt:,.2f}")

    return "\n".join(lines)

def format_week_summary(expenses):
    from datetime import timedelta
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    days = {}
    for e in expenses:
        days[e["date"]] = days.get(e["date"], 0) + float(e["amount"])

    lines = [f"📆 This Week — {week_start.strftime('%b %d')}–{today.strftime('%b %d, %Y')}\n"]
    total = 0
    for i in range(7):
        d = week_start + timedelta(days=i)
        amt = days.get(str(d), 0)
        total += amt
        day_name = d.strftime("%a %b %d")
        lines.append(f"{day_name}    ₱{amt:,.2f}")

    wallet_totals = {}
    for e in expenses:
        wkey = e.get("wallet_detail") or e["wallet"]
        wallet_totals[wkey] = wallet_totals.get(wkey, 0) + float(e["amount"])

    lines.append("─────────────────────────")
    lines.append(f"💸 Total:  ₱{total:,.2f}\n")
    for w, amt in wallet_totals.items():
        lines.append(f"{format_wallet_display(w)}:  ₱{amt:,.2f}")
    return "\n".join(lines)

def format_month_summary(expenses):
    from datetime import timedelta
    today = datetime.now()
    lines = [f"🗓️ This Month — {today.strftime('%B %Y')}\n"]

    weeks = {}
    for e in expenses:
        d = datetime.strptime(e["date"], "%Y-%m-%d")
        week_num = (d.day - 1) // 7 + 1
        weeks[week_num] = weeks.get(week_num, 0) + float(e["amount"])

    total = sum(float(e["amount"]) for e in expenses)
    for w, amt in sorted(weeks.items()):
        lines.append(f"Week {w}    ₱{amt:,.2f}")

    wallet_totals = {}
    for e in expenses:
        wkey = e.get("wallet_detail") or e["wallet"]
        wallet_totals[wkey] = wallet_totals.get(wkey, 0) + float(e["amount"])

    lines.append("─────────────────────────")
    lines.append(f"💸 Total:  ₱{total:,.2f}\n")
    for w, amt in wallet_totals.items():
        lines.append(f"{format_wallet_display(w)}:  ₱{amt:,.2f}")
    return "\n".join(lines)

# Export to Excel
def generate_excel(expenses, title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Header style
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Date", "Time", "Merchant", "Amount (₱)", "Paid Via", "Entry Type"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, e in enumerate(expenses, 2):
        wallet_display = e.get("wallet_detail") or e["wallet"]
        ws.cell(row=row, column=1, value=e["date"])
        ws.cell(row=row, column=2, value=e["time"][:5] if e["time"] else "")
        ws.cell(row=row, column=3, value=e["merchant"])
        ws.cell(row=row, column=4, value=float(e["amount"]))
        ws.cell(row=row, column=5, value=wallet_display)
        ws.cell(row=row, column=6, value=e["entry_type"])

    # Total row
    total_row = len(expenses) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=4, value=sum(float(e["amount"]) for e in expenses)).font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name

# Handlers
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    get_or_create_user(user.id, user.username, user.first_name)
    await update.message.reply_text(
        f"Hi {user.first_name}! 🎯 Know where your money goes, take control of your day.\n\n"
        f"Use the buttons below to get started:",
        reply_markup=MAIN_KEYBOARD
    )

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    user = update.effective_user
    get_or_create_user(user.id, user.username, user.first_name)

    if text == "➕ Add Expense":
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📸 Receipt Photo", callback_data="add_receipt"),
             InlineKeyboardButton("✏️ Manual Entry", callback_data="add_manual")],
            [InlineKeyboardButton("🔙 Back", callback_data="back_main")]
        ])
        await update.message.reply_text("How do you want to add it?", reply_markup=keyboard)

    elif text == "📊 My Spending":
        today_expenses = get_today_expenses(user.id)
        today_total = sum(float(e["amount"]) for e in today_expenses)
        await update.message.reply_text(
            f"📊 My Spending\n💸 Today's Total: ₱{today_total:,.2f}",
            reply_markup=get_spending_keyboard()
        )

    elif text == "📤 Export":
        await update.message.reply_text(
            "Export your expenses:",
            reply_markup=get_export_keyboard()
        )

async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user = query.from_user

    # Back to main
    if data == "back_main":
        await query.edit_message_text("Use the buttons below:", reply_markup=None)
        return ConversationHandler.END

    # Add expense
    if data == "add_receipt":
        if not check_limit(user.id):
            await query.edit_message_text(
                "⚠️ You've reached your 20 free expenses this month.\n\n"
                "Upgrade to Pro for ₱99/month — unlimited tracking!\n"
                "Contact @YourUsername to upgrade."
            )
            return ConversationHandler.END
        context.user_data["entry_type"] = "receipt"
        await query.edit_message_text("Send me your receipt photo! 📸")
        return WAITING_FOR_RECEIPT

    if data == "add_manual":
        if not check_limit(user.id):
            await query.edit_message_text(
                "⚠️ You've reached your 20 free expenses this month.\n\n"
                "Upgrade to Pro for ₱99/month — unlimited tracking!\n"
                "Contact @YourUsername to upgrade."
            )
            return ConversationHandler.END
        context.user_data["entry_type"] = "manual"
        await query.edit_message_text("How much did you spend?\n(e.g. 185)")
        return MANUAL_AMOUNT

    # Wallet selection
    if data.startswith("wallet_"):
        wallet = data.replace("wallet_", "")
        context.user_data["wallet"] = wallet
        if wallet == "CreditCard":
            await query.edit_message_text("Which credit card? (e.g. BPI, Metrobank)")
            return MANUAL_WALLET_DETAIL
        elif wallet == "OtherBanks":
            await query.edit_message_text("Which bank? (e.g. BDO, BPI, UnionBank)")
            return MANUAL_WALLET_DETAIL
        else:
            context.user_data["wallet_detail"] = None
            now = datetime.now()
            if "date" not in context.user_data:
                context.user_data["date"] = now.strftime("%Y-%m-%d")
            if "time" not in context.user_data:
                context.user_data["time"] = now.strftime("%H:%M:%S")
            await query.edit_message_text(
                format_expense_confirmation(context.user_data),
                reply_markup=get_confirm_keyboard()
            )
            return CONFIRM_EXPENSE

    # Confirm
    if data == "confirm_save":
        d = context.user_data
        save_expense(user.id, d["amount"], d["merchant"], d["wallet"], d.get("wallet_detail"), d["date"], d["time"], d["entry_type"])
        wallet_display = format_wallet_display(d["wallet"], d.get("wallet_detail"))
        await query.edit_message_text(f"✅ Saved! ₱{float(d['amount']):,.2f} at {d['merchant']} via {wallet_display}.")
        context.user_data.clear()
        return ConversationHandler.END

    if data == "confirm_cancel":
        await query.edit_message_text("❌ Cancelled. Nothing was saved.")
        context.user_data.clear()
        return ConversationHandler.END

    if data == "confirm_edit":
        await query.edit_message_text("What do you want to fix?", reply_markup=get_edit_keyboard())
        return EDIT_MENU

    if data == "back_confirm":
        await query.edit_message_text(
            format_expense_confirmation(context.user_data),
            reply_markup=get_confirm_keyboard()
        )
        return CONFIRM_EXPENSE

    # Edit fields
    if data == "edit_amount":
        await query.edit_message_text(f"Current amount: ₱{float(context.user_data['amount']):,.2f}\nType the correct amount:")
        return EDIT_AMOUNT

    if data == "edit_merchant":
        await query.edit_message_text(f"Current: {context.user_data['merchant']}\nType the correct merchant:")
        return EDIT_MERCHANT

    if data == "edit_wallet":
        await query.edit_message_text("Which wallet?", reply_markup=get_wallet_keyboard())
        return EDIT_WALLET

    if data == "edit_date":
        await query.edit_message_text(f"Current date: {context.user_data['date']}\nType correct date (e.g. 2026-06-16):")
        return EDIT_DATE

    # Spending views
    if data == "view_today":
        expenses = get_today_expenses(user.id)
        today = datetime.now().strftime("%B %d, %Y")
        msg = format_expense_list(expenses, f"📅 Today — {today}")
        await query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="back_spending")]]))

    if data == "view_week":
        expenses = get_week_expenses(user.id)
        msg = format_week_summary(expenses)
        await query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="back_spending")]]))

    if data == "view_month":
        expenses = get_month_expenses(user.id)
        msg = format_month_summary(expenses)
        await query.edit_message_text(msg, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="back_spending")]]))

    if data == "view_pick":
        await query.edit_message_text("Which date do you want to view?\nType it like this: June 10")
        return HISTORY_PICK_DATE

    if data == "back_spending":
        today_expenses = get_today_expenses(user.id)
        today_total = sum(float(e["amount"]) for e in today_expenses)
        await query.edit_message_text(
            f"📊 My Spending\n💸 Today's Total: ₱{today_total:,.2f}",
            reply_markup=get_spending_keyboard()
        )

    # Export
    if data in ["export_today", "export_week", "export_month"]:
        if data == "export_today":
            expenses = get_today_expenses(user.id)
            title = f"Today {datetime.now().strftime('%b %d %Y')}"
        elif data == "export_week":
            expenses = get_week_expenses(user.id)
            title = f"This Week"
        else:
            expenses = get_month_expenses(user.id)
            title = f"{datetime.now().strftime('%B %Y')}"

        if not expenses:
            await query.edit_message_text("No expenses to export for this period.")
            return

        await query.edit_message_text("Generating your Excel file... ⏳")
        filepath = generate_excel(expenses, title)
        with open(filepath, "rb") as f:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=f,
                filename=f"GastadorTracker - {title}.xlsx",
                caption=f"📊 Here's your export for {title}!"
            )
        os.unlink(filepath)

    if data == "export_pick":
        await query.edit_message_text("Which date do you want to export?\nType it like this: June 10")
        return EXPORT_PICK_DATE

async def handle_receipt_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Got it! Reading your receipt... ⏳")
    photo = update.message.photo[-1]
    file = await context.bot.get_file(photo.file_id)
    image_bytes = await file.download_as_bytearray()

    try:
        data = await read_receipt(bytes(image_bytes))
        context.user_data.update(data)
        await update.message.reply_text(
            format_expense_confirmation(context.user_data),
            reply_markup=get_confirm_keyboard()
        )
        return CONFIRM_EXPENSE
    except Exception as e:
        logger.error(f"Receipt read error: {e}")
        await update.message.reply_text(
            "Sorry, I couldn't read that receipt. Try Manual Entry instead.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✏️ Manual Entry", callback_data="add_manual"), InlineKeyboardButton("🔙 Back", callback_data="back_main")]])
        )
        return ConversationHandler.END

# Manual entry states
async def manual_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = float(update.message.text.replace(",", "").replace("₱", ""))
        context.user_data["amount"] = str(amount)
        await update.message.reply_text("Where did you spend it?\n(e.g. Jollibee, Mercury Drug)")
        return MANUAL_MERCHANT
    except:
        await update.message.reply_text("Please enter a valid amount (numbers only, e.g. 185):")
        return MANUAL_AMOUNT

async def manual_merchant(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["merchant"] = update.message.text
    await update.message.reply_text("Which wallet did you use?", reply_markup=get_wallet_keyboard(include_back=False))
    return MANUAL_WALLET

async def manual_wallet_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["wallet_detail"] = update.message.text
    now = datetime.now()
    if "date" not in context.user_data:
        context.user_data["date"] = now.strftime("%Y-%m-%d")
    if "time" not in context.user_data:
        context.user_data["time"] = now.strftime("%H:%M:%S")
    await update.message.reply_text(
        format_expense_confirmation(context.user_data),
        reply_markup=get_confirm_keyboard()
    )
    return CONFIRM_EXPENSE

# Edit states
async def edit_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = float(update.message.text.replace(",", "").replace("₱", ""))
        context.user_data["amount"] = str(amount)
        await update.message.reply_text(
            format_expense_confirmation(context.user_data),
            reply_markup=get_confirm_keyboard()
        )
        return CONFIRM_EXPENSE
    except:
        await update.message.reply_text("Please enter a valid amount:")
        return EDIT_AMOUNT

async def edit_merchant(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["merchant"] = update.message.text
    await update.message.reply_text(
        format_expense_confirmation(context.user_data),
        reply_markup=get_confirm_keyboard()
    )
    return CONFIRM_EXPENSE

async def edit_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        d = datetime.strptime(update.message.text.strip(), "%Y-%m-%d")
        context.user_data["date"] = d.strftime("%Y-%m-%d")
        await update.message.reply_text(
            format_expense_confirmation(context.user_data),
            reply_markup=get_confirm_keyboard()
        )
        return CONFIRM_EXPENSE
    except:
        await update.message.reply_text("Please use this format: 2026-06-16")
        return EDIT_DATE

async def history_pick_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text.strip()
        # Try parsing various formats
        for fmt in ["%B %d", "%b %d", "%Y-%m-%d", "%m/%d/%Y"]:
            try:
                d = datetime.strptime(text, fmt)
                if fmt in ["%B %d", "%b %d"]:
                    d = d.replace(year=datetime.now().year)
                date_str = d.strftime("%Y-%m-%d")
                expenses = get_date_expenses(update.effective_user.id, date_str)
                msg = format_expense_list(expenses, f"📅 {d.strftime('%B %d, %Y')}")
                await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="back_spending")]]))
                return ConversationHandler.END
            except:
                continue
        await update.message.reply_text("Couldn't read that date. Try: June 10 or 2026-06-10")
        return HISTORY_PICK_DATE
    except:
        await update.message.reply_text("Please try again. Example: June 10")
        return HISTORY_PICK_DATE

async def export_pick_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        text = update.message.text.strip()
        for fmt in ["%B %d", "%b %d", "%Y-%m-%d"]:
            try:
                d = datetime.strptime(text, fmt)
                if fmt in ["%B %d", "%b %d"]:
                    d = d.replace(year=datetime.now().year)
                date_str = d.strftime("%Y-%m-%d")
                expenses = get_date_expenses(update.effective_user.id, date_str)
                if not expenses:
                    await update.message.reply_text("No expenses found for that date.")
                    return ConversationHandler.END
                title = d.strftime("%B %d %Y")
                filepath = generate_excel(expenses, title)
                with open(filepath, "rb") as f:
                    await context.bot.send_document(
                        chat_id=update.message.chat_id,
                        document=f,
                        filename=f"GastadorTracker - {title}.xlsx",
                        caption=f"📊 Here's your export for {title}!"
                    )
                os.unlink(filepath)
                return ConversationHandler.END
            except:
                continue
        await update.message.reply_text("Couldn't read that date. Try: June 10")
        return EXPORT_PICK_DATE
    except:
        return EXPORT_PICK_DATE

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Cancelled.", reply_markup=MAIN_KEYBOARD)
    return ConversationHandler.END

# Callback data that should NOT start/continue a conversation — these are
# one-shot navigation/read actions handled by their own top-level handler.
NAV_PATTERN = (
    "^(back_main|view_today|view_week|view_month|back_spending|"
    "export_today|export_week|export_month)$"
)
# Callback data that STARTS a conversation flow (needs follow-up input).
ENTRY_PATTERN = "^(add_receipt|add_manual|view_pick|export_pick)$"

async def nav_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # User tapped a navigation button while a conversation was active:
    # run the action and end the conversation cleanly.
    context.user_data.clear()
    await handle_callback(update, context)
    return ConversationHandler.END

async def menu_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # User tapped a main-menu button mid-conversation: reset and show the menu.
    context.user_data.clear()
    await handle_message(update, context)
    return ConversationHandler.END

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[
            # Only the buttons that actually start a flow may begin a conversation.
            CallbackQueryHandler(handle_callback, pattern=ENTRY_PATTERN)
        ],
        states={
            WAITING_FOR_RECEIPT: [MessageHandler(filters.PHOTO, handle_receipt_photo)],
            CONFIRM_EXPENSE: [CallbackQueryHandler(handle_callback)],
            EDIT_MENU: [CallbackQueryHandler(handle_callback)],
            EDIT_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_amount)],
            EDIT_MERCHANT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_merchant)],
            EDIT_WALLET: [CallbackQueryHandler(handle_callback)],
            EDIT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_date)],
            MANUAL_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_amount)],
            MANUAL_MERCHANT: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_merchant)],
            MANUAL_WALLET: [CallbackQueryHandler(handle_callback)],
            MANUAL_WALLET_DETAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, manual_wallet_detail)],
            HISTORY_PICK_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, history_pick_date)],
            EXPORT_PICK_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, export_pick_date)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CommandHandler("start", start),
            # Let navigation / main-menu taps break out of a conversation cleanly.
            CallbackQueryHandler(nav_fallback, pattern=NAV_PATTERN),
            MessageHandler(filters.Regex("^(➕ Add Expense|📊 My Spending|📤 Export)$"), menu_fallback),
        ],
        per_message=False
    )

    app.add_handler(CommandHandler("start", start))
    # Conversation first, so its active text states receive input before the
    # generic text handler below can swallow it.
    app.add_handler(conv_handler)
    # One-shot navigation buttons, handled outside any conversation.
    app.add_handler(CallbackQueryHandler(handle_callback, pattern=NAV_PATTERN))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("GastadorTrackerBot is running!")
    app.run_polling()

if __name__ == "__main__":
    main()
